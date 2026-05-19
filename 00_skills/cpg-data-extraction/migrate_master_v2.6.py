"""
migrate_master_v2.6.py — v2.5.x (47열) → v2.6 (48열) 마이그레이션

용도:
  cpg-data-extraction v2.6의 신규 열 구조에 맞춰 기존 엑셀 파일을 48열로 변환한다.
  - 기본정보 시트 AU(notes, 47번째) 앞에 AU `analysis_set` 신열 삽입
  - 기존 notes는 AV로 한 칸 시프트
  - 모든 셀 서식(폰트·정렬·배경·테두리·number_format) 보존
  - 데이터행의 기존 notes 값 보존, 신규 AU(analysis_set)는 공란

적용 대상:
  - 마스터 파일: 90_Output/AF_CPG_data_extraction_<작업자>.xlsx
  - 추출 파일:   90_Output/extracts/AF_extract_*.xlsx
  - 예시 파일:   90_Output/AF_CPG_data_extraction_example*.xlsx

사용법:
  단일 파일:
    python migrate_master_v2.6.py path/to/file.xlsx
  여러 파일/와일드카드:
    python migrate_master_v2.6.py "90_Output/extracts/*.xlsx"
  디렉터리 (재귀 아님):
    python migrate_master_v2.6.py 90_Output/

안전장치:
  - 락 파일(`~$*.xlsx`) 존재 시 즉시 중단
  - 자동 백업: <원본>.v2.5.1.bak (같은 폴더)
  - 멱등성: 이미 48열이면 skip
  - 헤더 47열 정합 확인 (마지막 열명이 'notes'가 아니면 거부)
"""
import sys
import shutil
from copy import copy
from pathlib import Path
from glob import glob

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl이 설치되어 있지 않습니다. `pip install openpyxl` 후 재실행하세요.")
    sys.exit(1)

NEW_HEADER = "analysis_set"
EXPECTED_LAST_HEADER = "notes"
EXPECTED_COLS_BEFORE = 47
EXPECTED_COLS_AFTER = 48
SHEET_BASIC_HINT = "기본"  # 기본정보 시트 식별 힌트


def find_basic_sheet(wb):
    for name in wb.sheetnames:
        if SHEET_BASIC_HINT in name or name == "기본정보":
            return wb[name]
    # fallback: 첫 번째 시트
    return wb[wb.sheetnames[0]]


def has_lock_file(path: Path) -> bool:
    lock = path.parent / f"~${path.name}"
    return lock.exists()


def migrate_one(path: Path) -> str:
    """단일 파일 마이그레이션. 반환: 'OK' | 'SKIP_*' | 'ERROR_*'"""
    if not path.exists():
        return f"ERROR_NOT_FOUND"
    if path.name.startswith("~$"):
        return "SKIP_LOCK_FILE_INPUT"
    if has_lock_file(path):
        return "ERROR_LOCKED (Excel에서 열려있음 — 닫고 재실행)"

    try:
        wb = openpyxl.load_workbook(path)
    except Exception as e:
        return f"ERROR_LOAD: {e}"

    ws = find_basic_sheet(wb)
    n_col = ws.max_column

    # 멱등성 체크
    if n_col == EXPECTED_COLS_AFTER:
        col47 = ws.cell(row=1, column=47).value
        col48 = ws.cell(row=1, column=48).value
        if col47 == NEW_HEADER and col48 == EXPECTED_LAST_HEADER:
            wb.close()
            return "SKIP_ALREADY_V2.6"

    if n_col != EXPECTED_COLS_BEFORE:
        wb.close()
        return f"SKIP_UNEXPECTED_COL_COUNT ({n_col} cols, expected {EXPECTED_COLS_BEFORE})"

    last_header = ws.cell(row=1, column=47).value
    if last_header != EXPECTED_LAST_HEADER:
        wb.close()
        return f"SKIP_UNEXPECTED_LAST_HEADER ({last_header!r}, expected {EXPECTED_LAST_HEADER!r})"

    # 백업 (같은 폴더, 멱등)
    bkp = path.with_suffix(path.suffix + ".v2.5.1.bak")
    if not bkp.exists():
        shutil.copy2(path, bkp)

    # col 47의 모든 행 스타일 캡처
    template_per_row = {}
    for r in range(1, ws.max_row + 1):
        c = ws.cell(row=r, column=47)
        template_per_row[r] = {
            "font": copy(c.font),
            "fill": copy(c.fill),
            "border": copy(c.border),
            "alignment": copy(c.alignment),
            "number_format": c.number_format,
            "protection": copy(c.protection),
        }

    # col 47 -> col 48 시프트 (값 + 스타일)
    for r in range(1, ws.max_row + 1):
        src = ws.cell(row=r, column=47)
        dst = ws.cell(row=r, column=48, value=src.value)
        t = template_per_row[r]
        dst.font = copy(t["font"])
        dst.fill = copy(t["fill"])
        dst.border = copy(t["border"])
        dst.alignment = copy(t["alignment"])
        dst.number_format = t["number_format"]
        dst.protection = copy(t["protection"])

    # col 47에 analysis_set 신규 — 헤더만 값, 데이터행 공란, 스타일은 기존 그대로
    ws.cell(row=1, column=47, value=NEW_HEADER)
    for r in range(2, ws.max_row + 1):
        c = ws.cell(row=r, column=47, value=None)
        t = template_per_row[r]
        c.font = copy(t["font"])
        c.fill = copy(t["fill"])
        c.border = copy(t["border"])
        c.alignment = copy(t["alignment"])
        c.number_format = t["number_format"]
        c.protection = copy(t["protection"])

    # column_dimensions 시프트 (AU width 그대로 AV로)
    dims = ws.column_dimensions
    au_letter, av_letter = get_column_letter(47), get_column_letter(48)
    if au_letter in dims and dims[au_letter].width:
        au_w = dims[au_letter].width
        dims[av_letter].width = au_w
        dims[au_letter].width = 14  # analysis_set 적정 너비

    try:
        wb.save(path)
    except Exception as e:
        wb.close()
        return f"ERROR_SAVE: {e}"
    wb.close()

    # 사후 검증
    wb2 = openpyxl.load_workbook(path, read_only=True)
    ws2 = find_basic_sheet(wb2)
    ok = (
        ws2.max_column == EXPECTED_COLS_AFTER
        and ws2.cell(row=1, column=47).value == NEW_HEADER
        and ws2.cell(row=1, column=48).value == EXPECTED_LAST_HEADER
    )
    wb2.close()
    return "OK" if ok else "ERROR_POSTCHECK"


def expand_args(args):
    paths = []
    for a in args:
        p = Path(a)
        if p.is_dir():
            paths.extend(sorted(p.glob("*.xlsx")))
        elif "*" in a or "?" in a:
            paths.extend(Path(x) for x in sorted(glob(a)))
        else:
            paths.append(p)
    # 락 파일/임시파일 제거
    return [p for p in paths if not p.name.startswith("~$")]


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(0)
    targets = expand_args(sys.argv[1:])
    if not targets:
        print("대상 파일 없음.")
        sys.exit(1)
    print(f"마이그레이션 대상: {len(targets)}개 파일\n")
    results = {}
    for p in targets:
        r = migrate_one(p)
        results[str(p)] = r
        marker = "✓" if r == "OK" else ("·" if r.startswith("SKIP") else "✗")
        print(f"  {marker} {p.name}: {r}")
    print()
    ok = sum(1 for v in results.values() if v == "OK")
    sk = sum(1 for v in results.values() if v.startswith("SKIP"))
    er = sum(1 for v in results.values() if v.startswith("ERROR"))
    print(f"결과: 성공 {ok} / 건너뜀 {sk} / 실패 {er}")
    if er:
        sys.exit(2)


if __name__ == "__main__":
    main()
